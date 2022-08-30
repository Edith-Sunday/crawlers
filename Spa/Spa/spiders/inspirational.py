import scrapy
from openpyxl import load_workbook

from ..helpers import log_error
from ..items import InspirationalItem


class DuschbyggarnaSQARP(scrapy.Spider):
    name = "duschbyggarna_sqarp"
    error_log = []
    log_error = log_error

    def start_requests(self):
        yield scrapy.Request(
            "https://example.com",
            callback=self.parse_inspirational
        )

    def parse_inspirational(self, response):
        work_book = self.read_inspirational('inspirational.xlsx')
        ws = work_book['Inspirational Data']

        if ws.max_column > 4:
            print('There seems to be more columns for Inspirational data')

        for row in ws.iter_rows(min_row=5):

            Items = InspirationalItem()
            Items['entity_type'] = row[0].value
            Items['name'] = row[1].value
            Items['id'] = row[1].value
            # HE SAID TO CHECK THAT ID AND NAME ARE UNIQUE WHATS WITH THAT?
            Items['url'] = row[3].value
            Items['descriptions'] = {
                'EN': {
                    'Inspirational Description': {
                        'text': row[2].value,
                        'html': row[2].value,
                    }
                }
            }
            yield Items

    @staticmethod
    def read_inspirational(filename: str):
        wb = load_workbook(filename)
        return {sheet: wb[sheet] for sheet in wb.sheetnames}


    @staticmethod
    def check_sheets(filename: str):
        wb = load_workbook(filename)
        # this should print the name of the sheets in the workbook and compare to the sheetnames we have originally written
        sheet_names = ['Inspirational Data', 'Inspirational Images', 'Inspirational Videos']

        for name in wb.sheetnames:
            if name in sheet_names:
                continue
            else:
                print(f"{name} was added to workbook.")

    @staticmethod
    def read_image(ws):
    # ws = wb["Inspirational Images"]
        if ws.max_column > 6:
            # check that the max column is actually 6
            print('There seems to be more columns for Inspirational Images')

        for row in ws.iter_rows(min_row=5):
            Items = InspirationalItem()

            Items['image_data'] = {}

            image_url = row[3].value
            filename = row[4].value
            title = ''
            priority = row[2].value
            types = row[5].value

            if filename.endswith('.jpg'):
                filename = filename
            else:
                filename = title

            Items['image_data'][image_url] = {
                'type': types,
                'priority': priority,
                'filename': filename,
                'title': title,
            }
            yield Items

    read_image('inspirational.xlsx')

    @staticmethod
    def read_video(ws):
        # ws = wb["Inspirational Videos"]
        if ws.max_column > 6:
            # check that the max column is actually 6
            print('There seems to be more columns for Inspirational Videos')
        for row in ws.iter_rows(min_row=5):

            Items = InspirationalItem()
            Items['insp_video_data'] = {}

            video_url = row[2].value
            video_host = row[4].value
            video_type = row[5].value
            language = row[3].value

            Items['insp_video_data'][video_url] = {
                'video_host': video_host,
                'Language': language,
                'video_type': video_type,
            }
            yield Items

    read_video("inspirational.xlsx")
