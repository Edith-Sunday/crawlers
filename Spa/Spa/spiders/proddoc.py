from openpyxl import load_workbook

doc_info = self.read_doc

doc_id = doc_info.get(item['sku'])
if doc_id:
    continue
else:
    print(" document id not found")

def read_doc(filename: str):
    wb = load_workbook(filename)
    ws = wb["Product Documents"]

    file_data = {}

    for row in ws.iter_rows(min_row=5, max_row=-1):
        file_url = row[3].value
        language = row[4].value
        document_type = row[5].value

        file_data[file_url] = {
            'file_url': file_url,
            'Language': language,
            'document_type': document_type,
        }
    print(file_data)

    if ws.max_column > 6:
        # check that the max column is actually 6
        print('There seems to be more columns for Product Documents')