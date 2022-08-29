from openpyxl import load_workbook

image_info = self.read_images

article_id = image_info.get(item['sku'])
if article_id:
    continue
else:
    print("image id not found")


wb = load_workbook(filename)
ws = wb["Product Images"]

def read_images(filename: str):

    image_data = {}

    for row in ws.iter_rows(min_row=5, max_row=-1):
        image_url = row[4].value
        priority = row[3].value
        filename = row[5].value
        title = ''

        if filename.endswith('.jpg'):
            filename = filename
        else:
            filename = title

    #    NEED HELP HERE: if priority == "",it must set it to each image top priority to 1 and so on

        # thing is there are two images each for each article_id,thats where priority comes in
        # filename must not end in jpg,check title
        image_data[image_url] = {
                    'priority': priority,
                    'title': title,
                    'filename': filename,
                }
        print(image_data)

    if ws.max_column > 6:
        # check that the max column is actually 6
            print('There seems to be more columns for Product Images')