from openpyxl import load_workbook

video_info = self.read_video

_id = video_info.get(item['sku'])
if _id:
    continue
else:
    print("video id not found")

def read_video(filename: str):
    wb = load_workbook(filename)
    ws = wb["Product Videos"]

    video_data = {}

    for row in ws.iter_rows(min_row=5, max_row=-1):
        video_url = row[3].value
        video_host = row[4].value
        language = row[5].value
        video_type = row[6].value

        video_data[video_url] = {
            'video_host': video_host,
            'Language': language,
            'video_type': video_type,
        }
    print(video_data)

    if ws.max_column > 6:
        # check that the max column is actually 6
        print('There seems to be more columns for Product Videos')