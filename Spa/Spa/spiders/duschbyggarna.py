from openpyxl import load_workbook
from decimal import Decimal

supp_data = self.supplier_data

# matching
supplier_id = supp_data.get(item['sku'])
if supplier_id:
    continue
else:
    print("supplier_id not found")

supplier_name = supp_data.get(item['scraper'])
if supplier_name:
    continue
else:
    print("supplier_name not found")

product_url = supp_data.get(item['url'])
if product_url:
    continue
else:
    print("product_url not found")

    # match for the ones with ?
    # and throw warnings

@staticmethod
def supplier_data(filename: str):
    wb = load_workbook(filename)
    ws = wb["Supplier Data"]

    supplier_info = {}

    for row in ws.iter_rows(min_row=5, max_row=-1):

        supplier_name = row[3].value
        supplier_id = row[4].value
        product_title = row[5].value
        product_url = row[6].value
        order_package_unit = row[7].value
        list_price = Decimal(row[8].value)
        stock_status_max_delivery_time_business_days = Decimal(row[9].value)

            # Checks
        if stock_status_max_delivery_time_business_days == "":
            stock_status_max_delivery_time_business_days == -1

        if order_package_unit == 'st':
            order_package_unit == 1

        price = list_price * Decimal(1.25)
        if price != supp_data.get(item['rrp_value']):
            print("list_price not same as rrp_value")
        # does the above make sense?

        supplier_info ={
            "supplier_name": supplier_name,
            "supplier_id": supplier_id,
            "product_title": product_title,
            "product_url": product_url,
            "order_package_unit": order_package_unit,
            "list_price": list_price,
            "stock_status_max_delivery_time_business_days": stock_status_max_delivery_time_business_days,


        }

    if ws.max_column > 6:
        print('There seems to be more columns for Supplier Data')
        # ensure max column is actually 6

    return supplier_info


# Quick question: the ones labelled new field am i to create those fields in the items.py?

supplier_data('base.xlsx')
