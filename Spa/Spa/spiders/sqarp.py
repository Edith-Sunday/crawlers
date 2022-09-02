from datetime import datetime, timedelta
from decimal import Decimal
import scrapy
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..helpers import log_error
from ..items import CategoryItem, InspirationalItem, ProductItem


class DuschbyggarnaSQARP(scrapy.Spider):
    name = "duschbyggarna"
    error_log = []
    log_error = log_error
    urls = {}

    def start_requests(self):
        yield scrapy.Request(
            "https://example.com",
            callback=self.parse
        )

    def parse(self, response):
        base_work_book = self.read_file('data_input/duschbyggarna_sqarp_base.xlsx')
        inspiration_work_book = self.read_file('data_input/duschbyggarna_sqarp_inspirational.xlsx')
        self.check_sheets('data_input/duschbyggarna_sqarp_base.xlsx')

        if not self.is_valid(base_work_book['Document Overview']):
            print('Received invalid document')
            return

        ws = base_work_book['Product Base Data']
        packages = self.read_packages(base_work_book['Packages'])
        specs = self.read_specs(base_work_book['Product Attributes'])
        images = self.read_products_images(base_work_book['Product Images'])
        documents = self.read_products_docs(base_work_book['Product Documents'])
        videos = self.read_products_videos(base_work_book['Product Videos'])
        related_products = self.read_related_products(base_work_book['Related Products'])
        supplier_data = self.read_supplier_data(base_work_book['Supplier Data'])

        # inspiration
        inspiration_data = self.read_inspiration_data(inspiration_work_book['Inspirational Data'])
        inspiration_images = self.read_inspiration_images(inspiration_work_book['Inspirational Images'])
        inspiration_videos = self.read_inspiration_videos(inspiration_work_book['Inspirational Videos'])
        categories = {}

        variant_parents = {}

        if 'SEK' in ws.cell(4, 18).value:
            currency = 'SEK'
        else:
            currency = 'UNKNOWN'
            log_error(self, f'Could not extract currency from: {ws.cell(4, 18).value}')

        if 'incl' in ws.cell(4, 18).value:
            includes_taxes = True
        else:
            includes_taxes = False

        # for row in ws.iter_rows(min_row=5):
        for row_no in range(5, ws.max_row + 1):
            category_name = ws.cell(row_no, 7).value
            category_item = categories.get(category_name)
            if not category_item:
                category_item = CategoryItem()
                category_item['scraper'] = self.name
                category_item['item_type'] = 'category'
                category_item['title'] = category_name.lower()
                category_item['url'] = f'https://{category_name.lower()}'
                category_item['child_product_urls'] = []

            item = ProductItem()
            item['scraper'] = self.name
            item['item_type'] = 'product'

            sku = ws.cell(row_no, 2).value.strip()
            item['sku'] = sku
            # if sku not in ['9108CR', '34986', '34987', '35015', '34983', '34982', '34982RA', '34982DO', '34982SV', '34982BR']:
            #     continue

            # Rule for unique naming from SQARP is:
            # "SQARP Category" + "Brand" + "Series" + "Model" + "Name of variant group" + "Variant name"
            # However, using category and Brand in this seems wrong (from Ecombooster perspective)
            try:
                sqarp_category = ws.cell(row_no, 7).value.strip()
            except Exception:
                sqarp_category = ''
            try:
                sqarp_brand = ws.cell(row_no, 9).value.strip()
            except Exception:
                sqarp_brand = ''
            try:
                sqarp_series = ws.cell(row_no, 10).value.strip()
            except Exception:
                sqarp_series = ''
            try:
                sqarp_model = ws.cell(row_no, 12).value.strip()
            except Exception:
                sqarp_model = ''
            try:
                sqarp_variant_group = ws.cell(row_no, 14).value.strip()
            except Exception:
                sqarp_variant_group = ''
            try:
                sqarp_variant_name = ws.cell(row_no, 13).value.strip()
            except Exception:
                sqarp_variant_name = ''

            # title = f'{sqarp_category} - {sqarp_brand} - {sqarp_series} - {sqarp_model} - {sqarp_variant_group} - {sqarp_variant_name}'
            title = f'{sqarp_series} - {sqarp_model} - {sqarp_variant_group} - {sqarp_variant_name}'
            title = title.replace(' -  - ', ' - ').replace(' -  - ', ' - ').replace(' -  - ', ' - ').replace(' -  - ', ' - ')
            if title.startswith(' - '):
                title = title[3:]
            if title.endswith(' - '):
                title = title[:-3]

            item['title'] = title
            item['url'] = ws.cell(row_no, 4).value
            item['import_export_taric_code_eu'] = ws.cell(row_no, 5).value
            item['import_export_country_of_origin'] = ws.cell(row_no, 6).value

            product_type = ws.cell(row_no, 8).value
            if product_type == 'Single':
                item['product_type'] = 'SIMPLE'
            elif product_type == 'Variant':
                item['product_type'] = 'VARIANT_CHILD'
                item["variant_attributes"] = {}

                variant_parent_sku = ws.cell(row_no, 15).value
                item['variant_parent_sku'] = variant_parent_sku

                if variant_parent_sku in variant_parents:
                    variant_parent_item = variant_parents[variant_parent_sku]
                else:
                    variant_parent_item = ProductItem()
                    variant_parent_item['scraper'] = self.name
                    variant_parent_item['item_type'] = 'product'
                    variant_parent_item['sku'] = variant_parent_sku

                    variant_parent_title = f'{sqarp_series} - {sqarp_model} - {sqarp_variant_group}'
                    variant_parent_title = variant_parent_title.replace(' -  - ', ' - ').replace(' -  - ', ' - ')
                    if variant_parent_title.startswith(' - '):
                        variant_parent_title = variant_parent_title[3:]
                    if variant_parent_title.endswith(' - '):
                        variant_parent_title = variant_parent_title[:-3]

                    variant_parent_item['title'] = variant_parent_title
                    variant_parent_item['url'] = ws.cell(row_no, 4).value
                    variant_parent_item['import_export_taric_code_eu'] = ws.cell(row_no, 5).value
                    variant_parent_item['import_export_country_of_origin'] = ws.cell(row_no, 6).value
                    variant_parent_item['product_type'] = 'VARIANT_PARENT'
                    variant_parent_item['brand'] = ws.cell(row_no, 9).value
                    variant_parent_item['series'] = ws.cell(row_no, 10).value
                    variant_parent_item['unique_selling_points'] = ws.cell(row_no, 11).value
                    variant_parent_item['model'] = ws.cell(row_no, 12).value
                    variant_parent_item['product_descriptions'] = {
                        'EN': {
                            'Product Description': {
                                'text': ws.cell(row_no, 17).value,
                                'html': ws.cell(row_no, 17).value,
                            }
                        }
                    }
                    variant_parent_item['parent_category_url'] = category_item['url']
                    variant_parent_item["variant_attributes"] = {}
                    variant_parent_item['price_currency'] = 'SEK'

                variant_attributes = ws.cell(row_no, 16).value.split(',')
                variant_values = ws.cell(row_no, 13).value.split(',')
                if len(variant_attributes) != len(variant_values):
                    log_error(self, f'Failed to extract variant values for {item["sku"]}, length of '
                                    f'variant attributes did not match length of variant values!')
                else:
                    for attr_no in range(0, len(variant_attributes)):
                        variant_attribute = variant_attributes[attr_no].strip()
                        variant_value = variant_values[attr_no].strip()

                        item["variant_attributes"][variant_attribute] = variant_value

                        if variant_attribute not in variant_parent_item['variant_attributes']:
                            variant_parent_item['variant_attributes'][variant_attribute] = [variant_value, ]
                        elif variant_value not in variant_parent_item['variant_attributes'][variant_attribute]:
                            variant_parent_item['variant_attributes'][variant_attribute].append(variant_value)

                variant_parents[variant_parent_sku] = variant_parent_item

            else:
                log_error(self, f'Unknown product type: {product_type}')

            item['brand'] = ws.cell(row_no, 9).value
            item['series'] = ws.cell(row_no, 10).value
            item['unique_selling_points'] = ws.cell(row_no, 11).value
            item['model'] = ws.cell(row_no, 12).value
            # item['variant_name : variant_group_variables'] = f'{ws.cell(row_no, 12].value} : {ws.cell(row_no, 15].value}'
            item['attributes'] = {}
            item['attributes']["ecombooster_sku"] = ws.cell(row_no, 1).value
            item['attributes']["name_of_variant_group"] = ws.cell(row_no, 14).value
            item["attributes"]['variant_group_id'] = ws.cell(row_no, 15).value
            item['attributes'].update(packages.get(item['sku'], {}))

            item['attributes'].update({
                "specifications": specs.get(item['sku'], {})
            })

            if supplier_data.get(item['sku']):
                supplier = supplier_data.get(item['sku'])
                # Currently not used...
                # "order_package_unit": order_package_unit,
                # "list_price": list_price,

                item['order_package_quantity'] = supplier["order_package_quantity"]
                item['order_package_minimum_order_quantity'] = supplier["order_package_quantity"]

                stock_max_delivery_time_business_days = int(supplier['stock_max_delivery_time_business_days'])
                stock_max_delivery_time_days = 0
                for days in range(0, stock_max_delivery_time_business_days):
                    stock_max_delivery_time_days += 1
                    if days % 5 == 0:
                        stock_max_delivery_time_days += 2
                stock_status_eta = datetime.today() + timedelta(days=stock_max_delivery_time_days)
                stock_status_eta = stock_status_eta.strftime('%Y-%m-%d')
                item['stock_status_eta'] = stock_status_eta

            else:
                log_error(self, f'There was no supplier data for {item["sku"]}')

            item['product_descriptions'] = {
                'EN': {
                    'Product Description': {
                        'text': ws.cell(row_no, 17).value,
                        'html': ws.cell(row_no, 17).value,
                    }
                }
            }

            item['stock_status_refined'] = 'BACKORDER'

            item['price_currency'] = currency

            item['rrp_value'] = Decimal(ws.cell(row_no, 18).value)
            item['rrp_includes_taxes'] = includes_taxes

            ean_code = str(ws.cell(row_no, 19).value)
            item['part_numbers'] = {
                f'EAN_{ean_code}': {
                    'type': 'EAN',
                    'id': ean_code,
                }
            }

            item['parent_category_url'] = category_item['url']

            # media
            item['image_urls'] = images.get(item['sku'], {}).get('image_urls', [])
            item['image_data'] = images.get(item['sku'], {}).get('image_data', {})
            item['file_urls'] = documents.get(item['sku'], {}).get('file_urls', [])
            item['file_data'] = documents.get(item['sku'], {}).get('file_data', {})

            #  related products
            item['related_products'] = related_products.get(item['sku'], [])

            category_item['child_product_urls'].append(item['url'])
            categories[category_name] = category_item

            yield item

        # Yielding categories
        for name, category_item in categories.items():
            yield category_item

        # Yielding variant parents
        for sku, variant_parent_item in variant_parents.items():
            yield variant_parent_item

        for inspiration_item in inspiration_data:
            inspiration_item['image_urls'] = inspiration_images.get(inspiration_item['id']).get('image_urls')
            inspiration_item['image_data'] = inspiration_images.get(inspiration_item['id']).get('image_data')
            # inspiration_item['video_urls'] = inspiration_images.get(inspiration_item['id
            yield inspiration_item

    @staticmethod
    def read_file(filename: str):
        wb = load_workbook(filename)
        return {sheet: wb[sheet] for sheet in wb.sheetnames}

    @staticmethod
    def is_valid(worksheet) -> bool:
        # checks if the field exists after reading file
        export_config = {
            "Table format for relational data": "One relation per row",
            "Data on inspirational entities included": True,
            "Include html mark-up for descriptions": False,
            "Language": "sv",
            "Selected multi-value separator": ";"
        }

        input_config = {}

        for i in range(27, 33):
            e_value = f"E{i}"
            f_value = f"F{i}"

            if not worksheet[e_value].value and not worksheet[f_value].value:
                continue
            input_config.update({
                worksheet[e_value].value: worksheet[f_value].value,
            })

        return export_config == input_config

    def read_packages(self, ws: Worksheet):
        header_attr_mapping = {
            'Package Depth (mm)': 'package_depth_mm',
            'Package Height (mm)': 'package_height_mm',
            'Package Width (mm)': 'package_width_mm',
            'Package Volume (m3)': 'package_volume_m3',
            'Package Weight (kg)': 'package_weight_kg',
        }
        header_col_no_mapping = {}

        for col_no in range(4, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            # Debug
            # print(col_no, title)
            if title in header_attr_mapping:
                header_col_no_mapping[col_no] = title
            else:
                log_error(self, f'Found header {title} in sheet "Packages" that we do not manage!')

        attr = {}
        for row_no in range(5, ws.max_row + 1):
            sku = ws.cell(row_no, 1).value
            temp_attr = {}
            for col_no, header in header_col_no_mapping.items():
                value = ws.cell(row_no, col_no).value
                if value is not None:
                    temp_attr[header_attr_mapping[header]] = value

            attr[sku] = temp_attr

        return attr

    @staticmethod
    def read_specs(ws: Worksheet):
        attr = {}

        specs_columns = [cell.value for row in ws.iter_rows(min_row=3, max_row=3) for cell in row]
        title_columns = [cell.value for row in ws.iter_rows(min_row=4, max_row=4) for cell in row]

        for row in ws.iter_rows(min_row=5):
            specs = {}
            for idx, cell in enumerate(row):
                if specs_columns[idx] == 'Specifikationer':
                    value = cell.value
                    if value is not None:
                        # add cell.value to attributes
                        specs.update({title_columns[idx]: cell.value})
                        continue
            attr.update({row[1].value: specs})
        return attr

    def read_products_images(self, ws: Worksheet):
        images = {}
        cols = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 6:
            log_error(self, 'There seems to be more columns for Product Images')

        for row in ws.iter_rows(min_row=5):
            item_image_value = images.get(row[cols['Manufacturer Article ID']].value)
            if not item_image_value:
                item_image_value = {
                    'image_urls': [],
                    'image_data': {},
                    'priority': 0
                }

            image_url = row[cols['Image URL']].value
            filename = row[cols['Filename']].value
            title = ''

            if filename and filename.endswith('.jpg'):
                filename = filename
            else:
                filename = title

            item_image_value['image_urls'].append(image_url)
            item_image_value['image_data'][image_url] = {
                'priority': item_image_value['priority'] + 1,
                'title': title,
                'filename': filename,
            }
            item_image_value['priority'] = item_image_value['priority'] + 1

            images.update({
                row[cols['Manufacturer Article ID']].value: item_image_value
            })
        return images

    def read_products_videos(self, ws: Worksheet):
        videos = {}
        cols = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 7:
            log_error(self, 'There seems to be more columns for Product Videos')

        for row in ws.iter_rows(min_row=5):
            video_data = {}
            video_url = row[cols['Video URL']].value
            video_host = row[cols['Video Host']].value
            language = row[cols['Language']].value
            video_type = row[cols['Video Type']].value

            video_data[video_url] = {
                'video_host': video_host,
                'Language': language,
                'video_type': video_type,
            }
            videos.update({row[cols['Manufacturer Article ID']].value: video_data})
        return videos

    def read_products_docs(self, ws: Worksheet):
        documents = {}
        cols = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 6:
            log_error(self, 'There seems to be more columns for Product Documents')

        for row in ws.iter_rows(min_row=5):
            item_doc_value = documents.get(row[cols['Manufacturer Article ID']].value)
            if not item_doc_value:
                item_doc_value = {
                    'file_urls': [],
                    'file_data': {}
                }

            file_url = row[cols['Document URL']].value
            item_doc_value['file_urls'].append(
                file_url
            )
            item_doc_value['file_data'][file_url] = {
                'Language': row[cols['Language']].value,
                'document_type': row[cols['Document Type']].value,
            }

            documents.update({
                row[cols['Manufacturer Article ID']].value: item_doc_value
            })
        return documents

    def read_related_products(self, ws: Worksheet):
        related_products = {}
        cols = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 10:
            log_error(self, 'There seems to be more columns for Related Products')

        for row in ws.iter_rows(min_row=5):
            item_rel_data = related_products.get(row[cols['Manufacturer Article ID']].value)
            if not item_rel_data:
                item_rel_data = []

            quantity = row[cols['Quantity']].value
            if not quantity:
                quantity = 1

            relation_type = row[cols['Relation Type']].value
            if relation_type == 'addons':
                relation_type = 'OPTIONAL_PART'
            else:
                log_error(self, f'Found unhandled relation {relation_type} for SKU {item_rel_data}')

            item_rel_data.append({
                "sku": row[cols['Related Article ID']].value,
                "relation": relation_type,
                "quantity": quantity
            })

            related_products.update({
                row[cols['Manufacturer Article ID']].value: item_rel_data
            })
        return related_products

    def read_supplier_data(self, ws: Worksheet):
        supplier_data = {}
        cols = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 10:
            log_error(self, 'There seems to be more columns for Supplier Data')

        for row in ws.iter_rows(min_row=5):
            supplier_name = row[cols['Supplier Name']].value
            supplier_id = row[cols['Supplier Article ID']].value
            product_title = row[cols['Supplier Product Title']].value
            product_url = row[cols['Supplier Product URL']].value
            list_price = Decimal(row[cols['Supplier List Price (SEK, excl VAT)']].value)
            stock_status_max_delivery_time_business_days = row[cols['Supplier Max Delivery Time (Business Days)']].value

            if not stock_status_max_delivery_time_business_days:
                stock_status_max_delivery_time_business_days = 30

            order_package_unit = row[cols['Supplier Purchase Unit']].value
            order_package_quantity = -1
            if order_package_unit == 'st':
                order_package_quantity = 1
            elif order_package_unit == 'par':
                order_package_quantity = 2
            else:
                log_error(self, f'Found order_package_unit {order_package_unit} that was not handled for {supplier_id}')

            supplier_info = {
                # "supplier_name": supplier_name,
                # "supplier_id": supplier_id,
                # "product_title": product_title,
                # "product_url": product_url,
                "order_package_unit": order_package_unit,
                "order_package_quantity": order_package_quantity,
                "list_price": list_price,
                "stock_max_delivery_time_business_days": stock_status_max_delivery_time_business_days
            }
            supplier_data.update({
                row[cols['Manufacturer Article ID']].value: supplier_info
            })
        return supplier_data

    def check_sheets(self, filename: str):
        wb = load_workbook(filename)
        valid_sheet_names = ['Document Overview', 'Product Base Data', 'Packages', 'Product Attributes',
                             'Supplier Data', 'Product Images', 'Product Documents', 'Product Videos',
                             'Related Products', ]
        for name in wb.sheetnames:
            if name not in valid_sheet_names:
                log_error(self, f"Found Unknown Sheet {name} in workbook!")

    def read_inspiration_data(self, ws: Worksheet):
        data = []
        cols = {}
        unique_items = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 4:
            log_error(self, 'There seems to be more columns for Inspirational data')

        for row in ws.iter_rows(min_row=5):
            item = InspirationalItem()
            if unique_items.get(row[cols['Entity Name']].value):
                log_error(self, f'Found duplicate entity {row[cols["Entity Name"]].value}')

            item['scraper'] = self.name
            item['entity_type'] = row[cols['Entity Type']].value
            item['name'] = row[cols['Entity Name']].value
            item['id'] = row[cols['Entity Name']].value

            item['url'] = row[cols['Manufacturers Entity URL']].value
            item['descriptions'] = {
                'EN': {
                    'Inspirational Description': {
                        'text': row[cols['Description']].value,
                        'html': row[cols['Description']].value,
                    }
                }
            }
            data.append(item)
        return data

    def read_inspiration_images(self, ws: Worksheet):
        images = {}
        cols = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 6:
            log_error(self, 'There seems to be more columns for Inspirational Images')

        for row in ws.iter_rows(min_row=5):
            entity = images.get(row[cols['Entity Name']].value)
            if not entity:
                entity = {
                    "image_urls": [],
                    "image_data": {}
                }

            entity['image_urls'].append(row[cols['URL']].value)
            filename = row[cols['CDN Filename']].value
            title = ''
            priority = row[cols['Suggested Sorting']].value
            types = row[cols['Type']].value

            if filename and filename.endswith('.jpg'):
                filename = filename
            else:
                filename = title

            entity['image_data'][row[cols['URL']].value] = {
                'type': types,
                'priority': priority,
                'filename': filename,
                'title': title,
            }
            images.update({row[cols['Entity Name']].value: entity})
        return images

    def read_inspiration_videos(self, ws: Worksheet):
        videos = {}
        cols = {}

        for col_no in range(1, ws.max_column + 1):
            title = ws.cell(4, col_no).value
            cols[title] = col_no - 1

        if ws.max_column > 6:
            log_error(self, 'There seems to be more columns for Inspirational Images')

        for row in ws.iter_rows(min_row=5):
            entity = videos.get(row[cols['Entity Name']].value)
            if not entity:
                entity = {
                    "video_urls": [],
                    "video_data": {}
                }

            entity['video_urls'].append(row[cols['URL']].value)
            video_host = row[cols['Video Host']].value
            video_type = row[cols['Type']].value
            language = row[cols['Language']].value

            entity['video_data'][row[cols['URL']].value] = {
                'video_title': row[cols['Entity Name']].value,
                'video_host': video_host,
                'Language': language,
                'video_type': video_type,
            }
            videos.update({row[cols['Entity Name']].value: entity})
        return videos
