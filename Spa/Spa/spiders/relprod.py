from openpyxl import load_workbook

rel_prod = self.related_product

related_prod_id = rel_prod.get(item['sku'])
if related_prod_id:
    continue
else:
    print(" related product id not found")

def related_product(filename: str):
    wb = load_workbook(filename)
    ws = wb["Related Products"]

    rel_prod_data = {}

    for row in ws.iter_rows(min_row=5, max_row=-1):
        # this max_row=-1 doesnt seem to work o
        sku = row[6].value
        relation = row[4].value
        quantity = row[5].value

        if quantity == '':
            quantity == 1

        rel_prod_data = {
            "sku": sku,
            "relation": relation,
            "quantity": quantity,
        }
    print(rel_prod_data)
        #NEED HELP HERE he said something about a list whats that about?
        #  Also the print statement returns values for only one row, whats that about?

    if ws.max_column >10:
        # check that the max column is actually 10
        print('There seems to be more columns for Related Products')