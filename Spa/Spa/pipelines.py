import hashlib
import json
import re
# import requests
import scrapy

from datetime import date, datetime
from decimal import Decimal
from openpyxl import Workbook, load_workbook
# from requests.auth import HTTPBasicAuth

from scrapy.exceptions import DropItem

# useful for handling different item types with a single interface
from scrapy.pipelines.images import ImagesPipeline
from scrapy.pipelines.files import FilesPipeline
from itemadapter import ItemAdapter
import scrapy


class SpaPipeline:
    # TODO - Shall we use the pipeline here to create the output json, instead of using the built in function in
    #  scrapy to store the items to json... This approach will give us more control of the output... E.g.
    #  aggregating data so that we only need one (1) product item even though there was several URLs which
    #  showed the same product...

    output_filename = None
    meta_filename = None
    items_scraped = 0
    scraped_category_urls = []
    scraped_product_urls = []
    stock_statuses = []
    start_spider_timestamp = None

    items = {
        'categories': {},
        'products': {},
    }

    # List of URLs that should be skipped...
    skip_urls = []

    def open_spider(self, spider):
        """
        When the spider is opened, we will read meta data or if it does not exist, create an empty file.
        """
        spider.error_log = {}
        self.start_spider_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.meta_filename = f'scraped_output/{spider.name}_meta.json'
        # self.output_filename = f'scraped_output/{spider.name}_NEW_'\
        #                        f'{datetime.now().strftime("%Y-%m-%d_%H%M%S")}.json'

        if spider.name == 'hillerstorp':
            filename = 'hillerstorp-no-login.json'
            try:
                with open(filename, encoding='utf8') as f:
                    spider.products = json.load(f)
            except:
                error = f'Failed to open {filename}!'
                spider.log_error(error)

        try:
            with open(self.meta_filename, encoding='utf8') as f:
                self.items_meta = json.load(f)
        except Exception:
            self.items_meta = {
                'scraped_urls': {},
                'dead_skus': {},
            }

        # TODO - Go for integration with ecombooster via REST API...
        # api_url = "http://localhost:8000/api/v0/procurement/products"
        # response = requests.get(api_url, auth=HTTPBasicAuth('username', 'password'))
        # print(response.json())

    def close_spider(self, spider):
        """
        When the spider is closed, we will store meta data and send a report via email.
        """
        print(f'DONE! Scraped in total {self.items_scraped} URLs where of {len(self.items)} unique items...')

        if spider.name == 'brafab-no-login':
            filename = 'Brafab_artikelinfo_2022.xlsx'
            try:
                wb = load_workbook(filename)
                ws = wb.active
            except Exception:
                spider.error_log(f'FAILED to open the file {filename}, which is used to complement the scraped data!',
                                 severity='CRITICAL')
                return

            column_rrp = 'RRP SEK Inkl'
            column_rrp_no = -1
            column_rrp_discount = 'RRP Rabatt SEK Inkl'
            column_rrp_discount_no = -1
            ws_columns = {}
            for col_no in range(1, ws.max_column + 1):
                col_title = ws.cell(1, col_no).value.strip()
                if col_title == column_rrp:
                    column_rrp_no = col_no
                elif col_title == column_rrp_discount:
                    column_rrp_discount_no = col_no
                if column_rrp_no > 0 and column_rrp_discount_no > 0:
                    break

            if column_rrp_no == -1 and column_rrp_discount_no == -1:
                column_rrp_no = ws.max_column + 1
                ws.cell(1, column_rrp_no).value = column_rrp
                column_rrp_discount_no = ws.max_column + 1
                ws.cell(1, column_rrp_discount_no).value = column_rrp_discount
            elif column_rrp_no == -1:
                column_rrp_no = ws.max_column + 1
                ws.cell(1, column_rrp_no).value = column_rrp
            elif column_rrp_discount_no == -1:
                column_rrp_discount_no = ws.max_column + 1
                ws.cell(1, column_rrp_discount_no).value = column_rrp_discount

            warnings = []

            row_no = 2
            for row_no in range(2, ws.max_row + 1):
                sku = ws.cell(row_no, 1).value
                product = spider.products.pop(sku, None)
                if product is not None:
                    try:
                        rrp = Decimal(product['price_value'])
                        if rrp <= 0:
                            warnings.append([sku, ws.cell(row_no, 3).value, f'rrp was <= 0: {rrp}'])
                    except Exception:
                        rrp = str(product['price_value']).replace('.', ',')
                        warnings.append([sku, ws.cell(row_no, 3).value, f'rrp was not a number: {rrp}'])

                    try:
                        rrp_discounted = Decimal(product['price_discount_value'])
                        if rrp_discounted == '-1':
                            rrp_discounted = ''
                    except Exception:
                        rrp_discounted = str(product['price_discount_value']).replace('.', ',')

                    ws.cell(row_no, column_rrp_no).value = rrp
                    ws.cell(row_no, column_rrp_discount_no).value = rrp_discounted
                else:
                    warnings.append([sku, ws.cell(row_no, 3).value, 'was not found when scraping site'])

            for sku, product in spider.products.items():
                row_no += 1

                try:
                    rrp = Decimal(product['price_value'])
                    if rrp <= 0:
                        warnings.append([sku, ws.cell(row_no, 3).value, f'rrp was <= 0: {rrp}'])
                except Exception:
                    rrp = str(product['price_value']).replace('.', ',')
                    warnings.append([sku, ws.cell(row_no, 3).value, f'rrp was not a number: {rrp}'])

                try:
                    rrp_discounted = Decimal(product['price_discount_value'])
                    if rrp_discounted == '-1':
                        rrp_discounted = ''
                except Exception:
                    rrp_discounted = str(product['price_discount_value']).replace('.', ',')

                ws.cell(row_no, 1).value = sku
                ws.cell(row_no, 3).value = product['title']
                ws.cell(row_no, column_rrp_no).value = rrp
                ws.cell(row_no, column_rrp_discount_no).value = rrp_discounted

            wb.save(filename)

            for warning in warnings:
                spider.log_error(f'{str(warning[0]):<20} {str(warning[1]):<40} {str(warning[2])}', severity='WARNING')

        elif spider.name == 'hillerstorp-no-login':
            filename = 'hillerstorp-no-login.json'
            try:
                with open(filename, 'w', encoding='utf8') as f:
                    json.dump(spider.products, f)
            except:
                error = f'Failed to save {filename}!'
                spider.log_error(error)

        else:
            try:
                with open(self.meta_filename, 'w', encoding='utf8') as f:
                    json.dump(self.items_meta, f)
            except Exception:
                error = f'Failed to save {self.meta_filename}!'
                spider.log_error(error)

            # try:
            #     with open(self.output_filename, 'w', encoding='utf8') as f:
            #         json.dump(self.items, f)
            # except Exception:
            #     error = f'Failed to save {self.output_filename}!'
            #     spider.log_error(error)

        # TODO - Instead of saving the output as json, we should add the possibility to send all the data
        #  to ecombooster REST API, when all the data is scraped!

        self.send_report(spider, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    def send_report(self, spider, end_timestamp):
        from .helpers import fetch_email_to_list

        # using SendGrid's Python Library
        # https://github.com/sendgrid/sendgrid-python
        # https://app.sendgrid.com/guide/integrate/langs/python
        import base64
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition, \
            Personalization, Email

        html_content = f'<p>{spider.name} spindlades mellan<br>{self.start_spider_timestamp} och<br>{end_timestamp}.</p>' \
                       f'<p>Totalt hittades <b>{len(self.scraped_category_urls)}</b> kategori-länkar och ' \
                       f'<b>{len(self.scraped_product_urls)}</b> produkt-länkar.</p>'

        if spider.error_log:
            try:
                no_of_erros = 0
                error_content = '<ul>\n'
                for classification, severity_data in spider.error_log.items():
                    error_content += f'<li>{classification}\n'
                    error_content += '<ul>\n'
                    for severity, list_of_errors in severity_data.items():
                        no_of_erros += len(list_of_errors)
                        error_content += f'<li>{severity}\n'
                        error_content += '<ul>\n'
                        for error in list_of_errors:
                            error_content += f'<li>{error}</li>\n'
                        error_content += '</ul></li>\n'
                    error_content += '</ul></li>\n'
                error_content += '</ul>\n'
                html_content += f'<p>Vi hittade totalt <b>{no_of_erros}</b> fel/konstigheter!</p>\n'\
                                f'{error_content}'
            except Exception as err:
                html_content += f'<p>Något gick fel när listan med fel skulle skrivas ut! {err}</p>'
        else:
            html_content += f'<p>Vi hittade <b>INGA</b> fel/konstigheter! Jippi!</p>'

        to_list = Personalization()
        email_to_list = fetch_email_to_list()
        if not email_to_list:
            to_list.add_to(Email('tobias@ecombooster.io'))
            html_content = f'<p><b>COULD NOT FIND EMAIL ADDRESSES TO SEND THIS REPORT TO!</b></p>{html_content}'
        else:
            for email_address in email_to_list:
                to_list.add_to(Email(email_address))

        message = Mail(
            from_email='info@ecombooster.io',
            subject=f'Rapport spindling: Nilssons Möbler / {spider.name}',
            html_content=html_content)
        message.add_personalization(to_list)

        # with open('scrape_meta_frokenfraken_kundtjanst-diff_20211030_064757_vs_20211101_061749.xlsx', 'rb') as f:
        #     data = f.read()
        #     f.close()
        # encoded_file = base64.b64encode(data).decode()
        #
        # attachedFile = Attachment(
        #     FileContent(encoded_file),
        #     FileName('Report.xlsx'),
        #     FileType('application/xlsx'),
        #     Disposition('attachment')
        # )
        # message.attachment = attachedFile

        try:
            sg = SendGridAPIClient('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')
            response = sg.send(message)
            # print(response.status_code, response.body, response.headers)
        except Exception as e:
            print(e.message)

    def process_item(self, item, spider):
        """
        Starting point for processing items. Depending on what type of item we are dealing with, it will be sent
        to other functions to be dealt with.
        """
        self.items_scraped += 1

        url = item.get('url', None)
        item_type = item.get('item_type', None)

        if url is None or url.strip() == '':
            error = 'Dropping item since url was None or empty!'
            spider.log_error(error)
            DropItem(error)

        elif url in self.skip_urls:
            raise DropItem(f'Skipping URL {url}, since it was in the list of URLs to skip...')

        elif item_type == 'category':
            item = self.process_item_category(item, spider)

        elif item_type == 'product':
            item = self.process_item_product(item, spider)

        else:
            error = f'ERROR! Product at {url} had an unknown item_type "{item_type}"'
            spider.log_error(error)
            raise DropItem(error)

        if self.items_scraped % 50 == 0:
            # print(f'Scraped in total {self.items_scraped} URLs where of {len(self.)} unique items...')
            print(f'Scraped in total {self.items_scraped} URLs...')

        return item

    def process_item_category(self, item, spider):
        """
        This will process category items. Goal is to have an output that is streamlined and can be imported
        into ecombooster.io without any modifications to the json file that is produced by the scraper.
        """

        url = item['url']
        if url not in self.scraped_category_urls:
            self.scraped_category_urls.append(url)

        # TODO - decide if we need this approach. Would be good to use when we start using the REST API
        #  towards ecombooster... Store all items in a dict and then update in ecombooster at close()
        if url not in self.items['categories']:
            category_item = {
                'scraper': item['scraper'],
                'item_type': 'category',
                'id': item.get('id', None),
                'title': item['title'],
                'breadcrumbs': item.get('breadcrumbs', None),
                'parent_category_url': item.get('parent_category_url', None),
                'description_html': item.get('description_html', None),
                'description_excerpt_html': item.get('description_excerpt_html', None),
                'errors': [],
                'child_categories': [],
                'child_products': [],
                'related_categories': [],
                'image_urls': [],
                'image_details': [],
                'images': [],
            }
            self.items['categories'][url] = category_item
        else:
            category_item = self.items['categories'][url]

        # Let's loop over a bunch of lists in the item and make sure that we have saved all entries related
        # to this category...
        list_items = ['errors', 'child_categories', 'child_products', 'related_categories',
                      'image_urls', 'image_details', 'images', ]
        for list_item in list_items:
            for entry in item.get(list_item, []):
                if entry not in category_item[list_item]:
                    category_item[list_item].append(entry)

        self.items_meta['scraped_urls'][url] = {
            'item_type': 'category',
            'status_code': item.get('status_code', None),
            'categories': [],
            'products': [],
            'timestamp': item.get('timestamp', None),
        }

        # TODO - figure out if and how to calculate md5 hash. Which fields should be included and in which order.
        # md5 = self.calculate_md5hash(item)

        # TODO - figure out if we need a timestamp or not...
        # item['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return item

    def process_item_product(self, item, spider):
        """
        This will process product items. Goal is to have an output that is streamlined and can be imported
        into ecombooster.io without any modifications to the json file that is produced by the scraper.
        """
        # A bunch of lists with data used in the validations...
        list_of_product_types = ['SIMPLE', 'VARIANT_PARENT', 'VARIANT_CHILD', 'SIBLING']
        list_of_stock_statuses = ['IN_STOCK', 'BACKORDER', 'LAST_AVAILABLE', 'DISCONTINUED', 'BUILT_ON_ORDER', 'SPECIALORDER', 'COMING_SOON', 'NA', 'UNKNOWN', ]
        list_of_product_relationships = ['RELATED_PRODUCT', 'UP_SELL', 'HAS_PARTS', 'PART_OF',
                                         'RECOMMENDED_REPLACEMENT', 'OPTIONAL_PART', 'SIBLING']
        list_of_part_number_types = ['SKU', 'EAN', 'RSK', 'GTIN', 'UPC', ]

        list_of_currencies = ['EUR', 'USD', 'SEK', 'GPB', ]
        list_of_currency_symbols = ['€', '$', 'kr', '£']
        currency_symbol_to_iso_code = {
            '€': 'EUR',
            '$': 'USD',
            'kr': 'SEK',
            '£': 'GBP',
        }

        sku = item.get('sku', None)
        url = item.get('url', None)
        if url not in self.scraped_product_urls:
            self.scraped_product_urls.append(url)

        if sku is None or sku.strip() == '':
            error = f'Dropping item with url {url}! sku was None or empty "{sku}"'
            spider.log_error(error)
            DropItem(error)

        if url is None or url.strip() == '':
            error = f'Dropping item with sku {sku}! url was None or empty "{url}"'
            spider.log_error(error)
            DropItem(error)

        product_type = item.get('product_type', None)
        if product_type is None or product_type not in list_of_product_types:
            error = f'Dropping item with SKU {sku}! product_type was "{product_type}"'
            spider.log_error(error)
            DropItem(error)

        if product_type == 'VARIANT_CHILD' and item.get('variant_parent_sku', '').strip() == '':
            error = f'Dropping item with SKU {sku}! product_type was "{product_type}" but there was no ' \
                    f'variant_parent_sku defined!'
            spider.log_error(error)
            DropItem(error)

        if item.get('title', '').strip() == '':
            error = f'Dropping item with SKU {sku}! title was not defined!'
            spider.log_error(error)
            DropItem(error)

        if item.get('parent_category_url', '').strip() == '':
            error = f'WARNING! parent_category_url was not defined for {sku}'
            spider.log_error(error)

        stock_status_refined = item.get('stock_status_refined', '')
        if stock_status_refined not in list_of_stock_statuses:
            if product_type in ['SIMPLE', 'VARIANT_CHILD', ]:
                error = f'Dropping item with SKU {sku}! stock_status_refined was not set to accepted value: ' \
                        f'{stock_status_refined}'
                spider.log_error(error)
                DropItem(error)
            elif product_type in ['VARIANT_PARENT', ]:
                spider.logger.debug(f'{sku} did not have a correct stock_status "{stock_status_refined}", but is is a VARIANT_PARENT...')

        stock_status_eta = item.get('stock_status_eta', None)
        if stock_status_eta is not None:
            if product_type in ['SIMPLE', 'VARIANT_CHILD', ]:
                try:
                    item['stock_status_eta'] = date.fromisoformat(stock_status_eta)
                except Exception:
                    error = f'Dropping item with SKU {sku}! stock_status_eta did not have isoformat ' \
                            f'(YYYY-MM-DD), instead it was: "{stock_status_eta}"'
                    spider.log_error(error)
                    DropItem(error)

        product_descriptions = item.get('product_descriptions', {})
        for lang, descriptions in product_descriptions.items():
            if lang not in ['EN', 'FR', 'SV', 'NL']:
                error = f'Dropping item with SKU {sku}! product_description language "{lang}" is not supported!'
                spider.log_error(error)
                DropItem(error)
            for heading, content in descriptions.items():
                for text_type, text in content.items():
                    if text_type not in ['text', 'html']:
                        error = f'Dropping item with SKU {sku}! product_description text_type "{text_type}" is ' \
                                f'not supported!'
                        spider.log_error(error)
                        DropItem(error)

        # attributes = item.get('attributes', {})
        # for key, value in attributes.items():
        #     if type(value) is not dict:
        #         error = f'Dropping item with SKU {sku}! attributes / {key} is of ' \
        #                         f'not supported!'
        #                 spider.log_error(error)
        #                 DropItem(error)

        try:
            currency = item.get('price_currency', '').strip()
            if currency in list_of_currencies:
                pass
            if currency in list_of_currency_symbols:
                currency = currency_symbol_to_iso_code[currency]
                item['price_currency'] = currency
            if currency not in list_of_currencies:
                if product_type in ['SIMPLE', 'VARIANT_CHILD', ]:
                    error = f'Dropping item with SKU {sku}! price_currency was not set to accepted value: "{currency}"'
                    spider.log_error(error)
                    DropItem(error)
                elif product_type in ['VARIANT_PARENT', ]:
                    spider.logger.debug(f'{sku} did not have a correct currency "{currency}", but is is a '
                                        f'VARIANT_PARENT...')
        except BaseException:
            spider.log_error(f'Failed to extract currency at {item["url"]}: "{item.get("price_currency", "")}"')

        related_products = item.get('related_products', [])
        for relationship in related_products:
            if type(relationship) is not dict:
                error = f'Dropping item with SKU {sku}! related_products has an entry which is not a dict!'
                spider.log_error(error)
                DropItem(error)
            elif relationship.get('sku', '').strip() == '' and relationship.get('url', '').strip() == '':
                error = f'Dropping item with SKU {sku}! related_products has an entry which lacks both sku and url!'
                spider.log_error(error)
                DropItem(error)
            elif relationship.get('relation', '') not in list_of_product_relationships:
                error = f'Dropping item with SKU {sku}! related_products has an entry which has an unaccepted ' \
                        f'relation type: "{relationship.get("relation", "")}"'
                spider.log_error(error)
                DropItem(error)

        part_numbers = item.get('part_numbers', {})
        for key, value in part_numbers.items():
            if type(value) is not dict:
                error = f'Dropping item with SKU {sku}! part_numbers has an entry which is not a dict!'
                spider.log_error(error)
                DropItem(error)
            elif value.get('type', '').strip() not in list_of_part_number_types:
                error = f'Dropping item with SKU {sku}! part_numbers has an entry which has an unknown type: ' \
                        f'"{value.get("type", "")}"!'
                spider.log_error(error)
                DropItem(error)

        # TODO - decide if we need this approach. Would be good to use when we start using the REST API
        #  towards ecombooster... Store all items in a dict and then update in ecombooster at close()
        # if sku not in self.items['products']:
        #     product_item = {
        #         'scraper': item['scraper'],
        #         'item_type': 'product',
        #         'id': item.get('id', None),
        #         'title': item['title'],
        #         'description_html': item.get('description_html', None),
        #         'description_excerpt_html': item.get('description_excerpt_html', None),
        #         'urls': [url, ],
        #         'breadcrumbs': [item.get('breadcrumbs', None), ],
        #         'parent_category_urls': [item.get('parent_category_url', None), ],
        #         'errors': [],
        #         'child_categories': [],
        #         'child_products': [],
        #         'related_categories': [],
        #         'image_urls': [],
        #         'image_details': [],
        #         'images': [],
        #     }
        #     self.items['products'][sku] = product_item
        # else:
        #     product_item = self.items['products'][sku]
        #
        # list_items = ['errors', 'image_urls', 'image_details', 'images',
        #               'file_urls', 'file_details', 'files', 'video_urls', ]
        # for list_item in list_items:
        #     for entry in item.get(list_item, []):
        #         if entry not in product_item[list_item]:
        #             product_item[list_item].append(entry)
        #
        # self.items_meta['scraped_urls'][url] = {
        #     'item_type': 'product',
        #     'status_code': item.get('status_code', None),
        #     'sku': item.get('sku', None),
        #     'raw_stock_status': item.get('raw_stock_status', None),
        #     'timestamp': item.get('timestamp', None),
        # }

        # TODO - Decide which fields to use and in which order to calculate the md5 hash. It does not seem good
        #  enough to just take the whole item dict since different order of attributes will result in changed
        #  md5 hash...
        # md5 = self.calculate_md5hash(item)

        # TODO - figure out if we need a timestamp or not...
        # item['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        return item

    def calculate_md5hash(self, item):
        """
        This is used to calculate a md5-hash for the item.
        """
        # TODO In order for this to be useable, we need to define WHICH fields that shall be used in the calculation
        #  and in which order.
        # deepcopy is needed to avoid removing data in e.g. images from original item...
        # tmp_item = copy.deepcopy(item)
        # tmp_item.pop('timestamp', None)
        # try:
        #     for image in tmp_item['images']:
        #         image.pop('status', None)
        #         image.pop('checksum', None)
        # except Exception:
        #     pass
        #
        # item['md5_hash'] = hashlib.md5(str(tmp_item).encode('utf-8')).hexdigest()
        return


class MyImagesPipeline(ImagesPipeline):
    def get_media_requests(self, item, info):
        if self._Image.MAX_IMAGE_PIXELS:
            self._Image.MAX_IMAGE_PIXELS = None
        for image_url in item.get("image_urls", []):
            try:
                yield scrapy.Request(image_url)
            except Exception:
                pass

    def thumb_path(self, request, thumb_id, response=None, info=None):
        try:
            spider = info.spider.name
        except Exception:
            raise Exception('ERROR! Could not get name of spider in thumb_path()...')
        return f'{spider}/{super().thumb_path(request, thumb_id, response, info)}'

    def file_path(self, request, response=None, info=None, *, item=None):
        try:
            spider = info.spider.name
        except Exception:
            raise Exception('ERROR! Could not get name of spider in file_path()...')
        return f'{spider}/{super().file_path(request=request, response=response, info=info, item=item)}'


class MyFilesPipeline(FilesPipeline):
    def file_path(self, request, response=None, info=None, *, item=None):
        try:
            spider = info.spider.name
        except Exception:
            raise Exception('ERROR! Could not get name of spider in file_path()...')
        return f'{spider}/{super().file_path(request=request, response=response, info=info, item=item)}'


class MyLocalImagesPipeline(ImagesPipeline):
    def __init__(self, store_uri, download_func=None, settings=None):
        store_uri = settings.get('IMAGES_STORE_LOCAL')
        self._Image.MAX_IMAGE_PIXELS = None
        super().__init__(store_uri, download_func, settings)

    def thumb_path(self, request, thumb_id, response=None, info=None):
        try:
            spider = info.spider.name
        except Exception:
            raise Exception('ERROR! Could not get name of spider in thumb_path()...')
        return f'{spider}/{super().thumb_path(request, thumb_id, response, info)}'

    def file_path(self, request, response=None, info=None, *, item=None):
        try:
            spider = info.spider.name
        except Exception:
            raise Exception('ERROR! Could not get name of spider in file_path()...')
        return f'{spider}/{super().file_path(request=request, response=response, info=info, item=item)}'


class MyLocalFilesPipeline(FilesPipeline):
    def __init__(self, store_uri, download_func=None, settings=None):
        store_uri = settings.get('FILES_STORE_LOCAL')
        super().__init__(store_uri, download_func, settings)

    def file_path(self, request, response=None, info=None, *, item=None):
        try:
            spider = info.spider.name
        except Exception:
            raise Exception('ERROR! Could not get name of spider in file_path()...')
        return f'{spider}/{super().file_path(request=request, response=response, info=info, item=item)}'
